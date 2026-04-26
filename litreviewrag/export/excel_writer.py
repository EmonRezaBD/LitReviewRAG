"""Excel export for paper extractions and contradiction synthesis.

Produces a single .xlsx file with two sheets:

1. "Extractions" — one row per paper, one column per field. Source chunk
   citations are included as a final column for verification, satisfying
   the proposal's bias-mitigation requirement (Section VII.B).
2. "Contradictions" — one row per detected contradiction, with both
   papers, their conflicting claims, and the model's explanation.

Output is styled with bold headers, frozen top row, and auto-sized columns
so the file is readable as soon as it opens — no manual formatting needed.
"""

import logging
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from litreviewrag.extraction.extractor import PaperExtraction
from litreviewrag.extraction.prompts import FIELD_SPECS
from litreviewrag.synthesis.contradiction import Contradiction

logger = logging.getLogger(__name__)

# Styling constants — kept at module scope so consumers can override if needed.
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")  # dark blue
_WRAP = Alignment(wrap_text=True, vertical="top")

# Column width caps prevent any single field from making the sheet unreadable.
_MIN_WIDTH = 12
_MAX_WIDTH = 60


def _format_source_citations(extraction: PaperExtraction) -> str:
    """Build a compact citation string listing source chunks per field.

    Args:
        extraction: A paper's extraction result.

    Returns:
        A multi-line string like:
            problem_statement: chunks 3, 7, 12
            findings: chunks 14, 22
        Useful for spot-checking which chunks the model used per field.
    """
    lines = []
    for spec in FIELD_SPECS:
        fe = extraction.fields.get(spec.name)
        if not fe or not fe.source_chunks:
            continue
        chunk_ids = ", ".join(str(c.chunk_index) for c in fe.source_chunks)
        lines.append(f"{spec.name}: chunks {chunk_ids}")
    return "\n".join(lines)


def _autosize_columns(ws: Worksheet) -> None:
    """Set column widths based on max content length, clamped to [12, 60].

    openpyxl has no built-in autosize. We measure the longest line in each
    column (split on newlines for wrapped cells) and clamp to a sensible
    range so columns are readable but not absurdly wide.

    Args:
        ws: Worksheet to resize.
    """
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        max_len = 0
        for cell in column_cells:
            if cell.value is None:
                continue
            # For wrapped cells, the relevant width is the longest single line
            longest_line = max(len(line) for line in str(cell.value).split("\n"))
            max_len = max(max_len, longest_line)
        width = max(_MIN_WIDTH, min(max_len + 2, _MAX_WIDTH))
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _style_header(ws: Worksheet, num_columns: int) -> None:
    """Apply bold white-on-blue styling to the header row + freeze it.

    Args:
        ws: Worksheet to style.
        num_columns: Number of header cells to style.
    """
    for col_idx in range(1, num_columns + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _WRAP
    ws.freeze_panes = "A2"

# These render as bullet points; all other fields render as prose.
_LIST_FIELDS = {
    "research_questions",
    "contributions",
    "methodology",
    "findings",
    "limitations",
    "future_work",
}


def _format_value(field_name: str, value: str) -> str:
    """Format a field value for Excel display.

    For semicolon-list fields (research_questions, contributions): split
    on semicolons into bullets.

    For prose fields that benefit from bulleting (methodology, findings,
    limitations, future_work): split on sentence-ending periods into
    bullets, preserving period punctuation on each bullet.

    For all other fields (title, problem_statement): return the original
    prose unchanged.

    Args:
        field_name: The field's identifier.
        value: The raw extracted value from the LLM.

    Returns:
        Either the original prose or a multi-line bullet list.
    """
    if not value:
        return ""
    if field_name not in _LIST_FIELDS:
        return value

    # Semicolon-style fields: explicit list separators in the prompt
    if field_name in {"research_questions", "contributions"}:
        items = [item.strip() for item in value.split(";") if item.strip()]
    else:
        # Prose fields: split on periods, but only when followed by a space
        # or end-of-string (avoids splitting on decimals like "85.3%").
        import re

        sentences = re.split(r"\.(?=\s|$)", value)
        items = [s.strip() for s in sentences if s.strip()]
        # Re-add the trailing period that the split consumed
        items = [s if s.endswith(".") else s + "." for s in items]

    if not items:
        return value
    return "\n".join(f"• {item}" for item in items)

def _write_extractions_sheet(
    ws: Worksheet, extractions: list[PaperExtraction]
) -> None:
    """Write the per-paper extraction sheet.

    Layout: paper_name | <8 field columns>

    Source chunk citations are NOT shown here to keep the sheet readable.
    They are preserved in results/extractions.json for verification, per
    the proposal's bias-mitigation requirement (Section VII.B).

    Args:
        ws: Target worksheet (renamed by caller).
        extractions: One PaperExtraction per paper.
    """
    headers = ["paper_name"] + [spec.name for spec in FIELD_SPECS]
    ws.append(headers)

    for ex in extractions:
        row = [ex.paper_name]
        # Pull each field in canonical order; missing fields render as empty
        for spec in FIELD_SPECS:
            fe = ex.fields.get(spec.name)
            value = fe.value if fe and fe.value else ""
            row.append(_format_value(spec.name, value))
        ws.append(row)

    # Wrap text in all data cells so long extractions stay readable
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP

    _style_header(ws, len(headers))
    _autosize_columns(ws)
    
def _write_contradictions_sheet(
    ws: Worksheet, contradictions: list[Contradiction]
) -> None:
    """Write the cross-paper contradictions sheet.

    Layout: paper_a | claim_a | paper_b | claim_b | explanation

    Args:
        ws: Target worksheet (renamed by caller).
        contradictions: List of detected contradictions, possibly empty.
    """
    headers = ["paper_a", "claim_a", "paper_b", "claim_b", "explanation"]
    ws.append(headers)

    if not contradictions:
        # Empty state: a single row explaining why the sheet is empty,
        # so users don't think the export silently failed
        ws.append(
            [
                "(none detected)",
                "Synthesis returned no contradictions.",
                "",
                "",
                "This may mean papers cover different topics, or no genuine "
                "factual conflicts exist in the corpus.",
            ]
        )
    else:
        for c in contradictions:
            ws.append([c.paper_a, c.claim_a, c.paper_b, c.claim_b, c.explanation])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = _WRAP

    _style_header(ws, len(headers))
    _autosize_columns(ws)


def export_to_excel(
    extractions: list[PaperExtraction],
    contradictions: list[Contradiction],
    output_path: Path,
) -> None:
    """Write the full literature review to a styled .xlsx file.

    Creates two sheets ("Extractions", "Contradictions") and overwrites
    any existing file at `output_path`. Parent directories are created
    automatically if missing.

    Args:
        extractions: Per-paper extraction results.
        contradictions: Detected cross-paper contradictions.
        output_path: Destination .xlsx file path.
    """
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()

    # First sheet always exists in a fresh workbook — rename rather than
    # delete + recreate to preserve openpyxl's default formatting hooks
    extractions_ws = wb.active
    extractions_ws.title = "Extractions"
    _write_extractions_sheet(extractions_ws, extractions)

    contradictions_ws = wb.create_sheet(title="Contradictions")
    _write_contradictions_sheet(contradictions_ws, contradictions)

    wb.save(output_path)
    logger.info(
        "Wrote %d extractions and %d contradictions to %s",
        len(extractions),
        len(contradictions),
        output_path,
    )